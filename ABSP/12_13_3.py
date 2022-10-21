import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('12_13_1.xlsx')
# 翻转电子表格中行和列的单元格
sheet = wb.active
sheetData = []
for row in range(1, sheet.max_row+1):
    sheetData.append([])
    for col in range(1, sheet.max_column+1):
        sheetData[-1].append(sheet[get_column_letter(col)+str(row)].value)
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = sheet.title
for row in range(len(sheetData)):
    for col in range(len(sheetData[0])):
        new_sheet[get_column_letter(row+1)+str(col+1)] = sheetData[row][col]
new_wb.save('update.xlsx')
