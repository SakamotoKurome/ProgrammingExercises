import openpyxl
import os
from openpyxl.utils import get_column_letter
# 一行一个单元格，一列一个文本文件
# 假定所有要处理的文本文件在某个目录中
jobs_dir = os.path.join(os.getcwd(), 'jobs')
files = os.listdir(jobs_dir)
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(len(files)):
    with open(os.path.join(jobs_dir, files[i])) as f:
        lines = f.readlines()
    for j in range(len(lines)):
        sheet[get_column_letter(i+1)+str(j+1)] = lines[j].strip()
wb.save('jobs.xlsx')
