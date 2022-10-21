import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
# 从命令行接受数字 N
N = int(input('Input a number: '))
# 在一个 Excel 电子表格中创建一个 N×N 的乘法表
wb = openpyxl.Workbook()
sheet = wb.active
for row in range(1, N+2):
    for col in range(1, N+2):
        if col == 1 and row == 1:
            continue
        elif row == 1:
            sheet[get_column_letter(col)+'1'] = col - 1
        elif col == 1:
            sheet['A'+str(row)] = row - 1
        else:
            sheet[get_column_letter(col)+str(row)] = (col-1) * (row-1)
# 行 1 和列 A 应该用做标签,应该使用粗体。
col = sheet.column_dimensions['A']
col.font = Font(bold=True)
row = sheet.row_dimensions[1]
row.font = Font(bold=True)
wb.save('12_13_1.xlsx')
