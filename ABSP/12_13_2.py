import openpyxl
from openpyxl.utils import get_column_letter
import sys
# 接受两个整数和一个文件名字符串
# python blankRowInserter.py 3 2 myProduce.xlsx
file = sys.argv[3]
N = int(sys.argv[1])
M = int(sys.argv[2])
# 读入电子表格的内容
wb = openpyxl.load_workbook(file)
sheet = wb.active
# 程序应该从第 N 行开始,在电子表格中插入 M 个空行。
# 从N行开始，将所有的单元格向后移 M 行
for row in range(sheet.max_row+1, N-1, -1):
    for col in range(sheet.max_column, 0, -1):
        sheet[get_column_letter(col)+str(row+M)] =\
            sheet[get_column_letter(col)+str(row)].value
# 将空行置空
for row in range(N, N+M):
    for col in range(1, sheet.max_column+1):
        sheet[get_column_letter(col)+str(row)] = None
wb.save('update.xlsx')
